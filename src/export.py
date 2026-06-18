import io
from datetime import datetime

import pandas as pd
from fpdf import FPDF
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font
from PIL import Image as PILImage

from src.charts import build_commodity_chart, build_geo_choropleth

CHART_PNG_DIMS = (800, 320)  # px, used for commodity line charts in exports
MAP_PNG_DIMS = (800, 440)  # px, used for the sourcing choropleth in exports


def _fig_to_png(fig, width, height):
    """Returns None instead of raising if kaleido can't launch a browser to
    render with - environments like Streamlit Cloud don't always have the
    system Chrome dependencies available, and a chart shouldn't break the
    whole export when that happens."""
    try:
        return fig.to_image(format="png", width=width, height=height, scale=2)
    except Exception:
        return None

RISK_COLOR_RGB = {
    "Low Risk": (46, 204, 113),
    "Moderate Risk": (243, 156, 18),
    "High Risk": (230, 126, 34),
    "Critical Risk": (231, 76, 60),
}

CATEGORY_LABELS = {
    "supplier": "Supplier Concentration",
    "commodity": "Commodity Price",
    "logistics": "Logistics & Shipping",
    "geopolitical": "Geopolitical",
    "regulatory": "Regulatory & Trade",
}


_UNICODE_REPLACEMENTS = {
    "—": "-", "–": "-",  # em dash, en dash
    "‘": "'", "’": "'",  # curly single quotes
    "“": '"', "”": '"',  # curly double quotes
    "…": "...",  # ellipsis
    "•": "-",  # bullet
    " ": " ",  # non-breaking space
}


def _pdf_safe(text):
    """Core Helvetica only supports latin-1; AI-generated text often has
    em-dashes/smart quotes that would otherwise crash fpdf2 mid-render."""
    text = str(text)
    for char, replacement in _UNICODE_REPLACEMENTS.items():
        text = text.replace(char, replacement)
    return text.encode("latin-1", "replace").decode("latin-1")


def _risk_label(score):
    if score <= 30:
        return "Low Risk"
    elif score <= 60:
        return "Moderate Risk"
    elif score <= 80:
        return "High Risk"
    return "Critical Risk"


def _commodity_rows(commodity_data):
    rows = []
    for name, history in commodity_data.items():
        if len(history) >= 2:
            latest = history[-1]["value"]
            pct_change = (history[-1]["value"] - history[0]["value"]) / history[0]["value"] * 100
            rows.append({
                "Commodity": name,
                "Latest Value": round(latest, 2),
                "% Change (period)": round(pct_change, 1),
            })
        else:
            rows.append({"Commodity": name, "Latest Value": "N/A", "% Change (period)": "N/A"})
    return rows


def _shipping_rows(shipping_status, logistics_result):
    rows = []
    for route_name, route_data in shipping_status.items():
        route_score = logistics_result["by_route"][route_name]
        rows.append({
            "Route": route_name,
            "Status": route_data["status"],
            "Delay (days)": route_data["delay_days"],
            "Cost Premium (%)": route_data["cost_premium_pct"],
            "Risk Score": route_score["final"],
            "News Signal": f"+{route_score['alert_adjustment']}" if route_score["alert_adjustment"] > 0 else "None",
        })
    return rows


def _sourcing_rows(by_country):
    rows = []
    for code, data in sorted(by_country.items(), key=lambda x: x[1]["weight"], reverse=True):
        rows.append({
            "Country": data["name"],
            "Code": code,
            "% of Sourcing": round(data["weight"] * 100, 1),
            "Product": data["product"],
            "Risk Score": data["final"],
        })
    return rows


def _compliance_status_text(compliance_results):
    flagged = [(name, r) for name, r in compliance_results if r["sanctioned"]]
    any_checked = any(r["checked"] for _, r in compliance_results)
    if flagged:
        return "; ".join(f"{name} - potential match: {r['matched_name']}" for name, r in flagged)
    if any_checked:
        return "Clear - no name match found on the US Consolidated Screening List."
    if compliance_results:
        return "Not checked - compliance screening API key not configured."
    return "Not checked - no company/supplier name available."


def _supplier_risk_rows(result, single_source_dependency, compliance_results):
    supplier_score = result["sub_scores"]["supplier"]
    dep_label, dep_pct, dep_country = single_source_dependency
    dependency_text = (
        f"{dep_label} - {dep_pct:.0f}% sourced from {dep_country}" if dep_country else "Unknown - no sourcing data"
    )
    return [
        {"Metric": "Supplier Risk Rating", "Value": f"{supplier_score} ({_risk_label(supplier_score)})"},
        {"Metric": "Single Source Dependency", "Value": dependency_text},
        {"Metric": "Supplier Compliance Status", "Value": _compliance_status_text(compliance_results)},
    ]


def _logistics_risk_rows(result, shipment_delays, port_congestion, on_time_rate):
    avg_delay_days, delay_label = shipment_delays
    port_score, port_routes = port_congestion
    transportation_risk_index = result["sub_scores"]["logistics"]
    port_text = (
        f"{port_score:.1f} ({_risk_label(port_score)})" if port_routes else "No tracked port routes available"
    )
    return [
        {"Metric": "Shipment Delays", "Value": f"{avg_delay_days:.1f} days average ({delay_label})"},
        {"Metric": "Port Congestion", "Value": port_text},
        {
            "Metric": "Transportation Risk Index",
            "Value": f"{transportation_risk_index} ({_risk_label(transportation_risk_index)})",
        },
        {
            "Metric": "On-Time Delivery Rate",
            "Value": f"{on_time_rate:.1f}% (estimate derived from Transportation Risk Index)",
        },
    ]


def _alert_band_text(alert_name, country_name, alert):
    if not country_name:
        return f"{alert_name}: No sourcing data available to check."
    adjustment = alert["adjustment"]
    if adjustment >= 12:
        band = "High"
    elif adjustment >= 6:
        band = "Elevated"
    elif adjustment > 0:
        band = "Watch"
    else:
        band = "Normal"
    return (
        f"{band} - {country_name} (top sourcing country): {alert['recent_count']} recent mentions "
        f"vs {alert['baseline_weekly_rate']}/week normal."
    )


def _geo_external_rows(top_country_name, disaster_alert, pol_reg_score, weather_alert, conflict_alert):
    return [
        {"Metric": "Natural Disaster Alerts", "Value": _alert_band_text("Natural Disaster Alerts", top_country_name, disaster_alert)},
        {"Metric": "Political/Regulatory Risks", "Value": f"{pol_reg_score} ({_risk_label(pol_reg_score)})"},
        {"Metric": "Weather Impact", "Value": _alert_band_text("Weather Impact", top_country_name, weather_alert)},
        {"Metric": "Regional Conflict Alerts", "Value": _alert_band_text("Regional Conflict Alerts", top_country_name, conflict_alert)},
    ]


def generate_excel_report(
    industry, company_name, time_horizon, result, ai_summary, recommendations,
    commodity_data, shipping_status, logistics_result, by_country,
    critical_alerts, high_risk_suppliers, disruption_band,
    single_source_dependency, compliance_results,
    shipment_delays, port_congestion, on_time_rate,
    top_country_name, disaster_alert, pol_reg_score, weather_alert, conflict_alert,
):
    """Returns the .xlsx file as bytes, ready for st.download_button."""
    overview_df = pd.DataFrame([
        {"Field": "Industry", "Value": industry},
        {"Field": "Company", "Value": company_name or "(industry-level, no company specified)"},
        {"Field": "Time Horizon", "Value": time_horizon},
        {"Field": "Report Generated", "Value": datetime.now().strftime("%Y-%m-%d %H:%M")},
        {"Field": "Overall Risk Score", "Value": result["total"]},
        {"Field": "Overall Risk Level", "Value": result["label"]},
    ])

    alerts_df = pd.DataFrame(
        [{"Category": label, "Score": score, "Risk Level": _risk_label(score)} for label, score in critical_alerts]
        or [{"Category": "None", "Score": "", "Risk Level": ""}]
    )

    suppliers_df = pd.DataFrame(
        [
            {"Name": s["name"], "Detail": s["detail"], "Risk Score": s["risk"], "Risk Level": s["label"]}
            for s in high_risk_suppliers
        ]
        or [{"Name": "None", "Detail": "", "Risk Score": "", "Risk Level": ""}]
    )

    disruption_df = pd.DataFrame([{"Value": disruption_band[0]}])

    exec_score_df = pd.DataFrame([
        {"Metric": "Overall Risk Score", "Value": f"{result['total']} / 100"},
        {"Metric": "Overall Risk Level", "Value": result["label"]},
    ])

    scores_df = pd.DataFrame([
        {
            "Category": CATEGORY_LABELS.get(key, key),
            "Score": value,
            "Risk Level": _risk_label(value),
            "Weight": f"{result['weights'][key]:.0%}",
        }
        for key, value in result["sub_scores"].items()
    ])

    summary_df = pd.DataFrame([{"AI Risk Brief": ai_summary}])

    recs_df = pd.DataFrame([
        {"#": i, "Recommendation": rec["title"], "Detail": rec["detail"], "Priority": rec["priority"]}
        for i, rec in enumerate(recommendations, 1)
    ])

    commodity_df = pd.DataFrame(_commodity_rows(commodity_data))
    shipping_df = pd.DataFrame(_shipping_rows(shipping_status, logistics_result))
    sourcing_df = pd.DataFrame(_sourcing_rows(by_country))
    supplier_risk_df = pd.DataFrame(
        _supplier_risk_rows(result, single_source_dependency, compliance_results)
    )
    logistics_risk_df = pd.DataFrame(
        _logistics_risk_rows(result, shipment_delays, port_congestion, on_time_rate)
    )
    geo_external_df = pd.DataFrame(
        _geo_external_rows(top_country_name, disaster_alert, pol_reg_score, weather_alert, conflict_alert)
    )

    sheets = [
        ("Overview", overview_df),
        ("Risk Scores", scores_df),
        ("Supplier Risk", supplier_risk_df),
        ("Logistics Risk", logistics_risk_df),
        ("Geographic & External Risk", geo_external_df),
        ("AI Summary", summary_df),
        ("Recommendations", recs_df),
        ("Commodity Prices", commodity_df),
        ("Shipping Routes", shipping_df),
        ("Sourcing Breakdown", sourcing_df),
    ]

    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        # Executive Summary stacks several small tables on one sheet (rather than
        # giving each its own tab), with bold section titles written directly via
        # openpyxl since pandas' to_excel has no header-styling option of its own.
        exec_sheet_name = "Executive Summary"
        row_cursor = 0
        exec_score_df.to_excel(writer, sheet_name=exec_sheet_name, index=False, startrow=row_cursor)
        row_cursor += len(exec_score_df) + 1 + 2
        exec_ws = writer.sheets[exec_sheet_name]

        for title, df in [
            ("Critical Risk Alerts", alerts_df),
            ("High-Risk Suppliers", suppliers_df),
            ("Disruption Probability", disruption_df),
        ]:
            exec_ws.cell(row=row_cursor + 1, column=1, value=title).font = Font(bold=True)
            row_cursor += 1
            df.to_excel(writer, sheet_name=exec_sheet_name, index=False, startrow=row_cursor)
            row_cursor += len(df) + 1 + 2

        for col_letter, width in [("A", 30), ("B", 40), ("C", 16), ("D", 16)]:
            exec_ws.column_dimensions[col_letter].width = width

        for sheet_name, df in sheets:
            df.to_excel(writer, sheet_name=sheet_name, index=False)

        # Auto-fit column widths roughly to content length, since the default is too narrow.
        for sheet_name, df in sheets:
            worksheet = writer.sheets[sheet_name]
            for col_idx, col_name in enumerate(df.columns, 1):
                max_len = max([len(str(col_name))] + [len(str(v)) for v in df[col_name]])
                worksheet.column_dimensions[worksheet.cell(row=1, column=col_idx).column_letter].width = min(
                    max(max_len + 2, 12), 80
                )

        # Charts/map are appended as images below each sheet's data table -
        # openpyxl can't embed a live Plotly figure, only a rendered bitmap.
        commodity_sheet = writer.sheets["Commodity Prices"]
        row_cursor = len(commodity_df) + 3
        width_px, height_px = CHART_PNG_DIMS
        for name, history in commodity_data.items():
            if len(history) < 2:
                continue
            png_bytes = _fig_to_png(build_commodity_chart(name, history), width_px, height_px)
            if png_bytes is None:
                continue
            xl_image = XLImage(PILImage.open(io.BytesIO(png_bytes)))
            xl_image.width, xl_image.height = width_px // 2, height_px // 2
            commodity_sheet.add_image(xl_image, f"A{row_cursor}")
            row_cursor += (height_px // 2) // 19 + 2  # rows are ~19px tall by default

        sourcing_sheet = writer.sheets["Sourcing Breakdown"]
        map_width_px, map_height_px = MAP_PNG_DIMS
        map_png_bytes = _fig_to_png(build_geo_choropleth(by_country), map_width_px, map_height_px)
        if map_png_bytes is not None:
            map_xl_image = XLImage(PILImage.open(io.BytesIO(map_png_bytes)))
            map_xl_image.width, map_xl_image.height = map_width_px // 2, map_height_px // 2
            sourcing_sheet.add_image(map_xl_image, f"A{len(sourcing_df) + 3}")

    return buffer.getvalue()


def _fit_cell_text(pdf, text, width):
    """Truncates with an ellipsis if text is wider than its column - fpdf2's
    cell() doesn't wrap or clip, it just overflows visually into the next cell."""
    text = str(text)
    max_width = width - 2  # small padding so text doesn't touch the border
    if pdf.get_string_width(text) <= max_width:
        return text
    while text and pdf.get_string_width(text + "...") > max_width:
        text = text[:-1]
    return (text + "...") if text else "..."


def _place_image(pdf, epw, png_bytes, width_px, height_px, margin_bottom=15):
    """Places an image at full content width, breaking to a new page first if
    it wouldn't fit - fpdf2's image() doesn't auto-paginate like cell() does.
    No-ops if png_bytes is None (chart failed to render, e.g. no headless
    Chrome available in this environment)."""
    if png_bytes is None:
        return
    height_mm = epw * (height_px / width_px)
    if pdf.get_y() + height_mm > pdf.h - margin_bottom:
        pdf.add_page()
    pdf.set_x(10)
    pdf.image(io.BytesIO(png_bytes), x=10, y=pdf.get_y(), w=epw, h=height_mm)
    pdf.set_xy(10, pdf.get_y() + height_mm + 4)


def _pdf_table(pdf, epw, headers, rows, col_weights):
    """Draws a simple bordered table. col_weights are relative widths summing
    to 1.0; converted to mm against the effective page width."""
    col_widths = [epw * w for w in col_weights]

    pdf.set_font("Helvetica", "B", 9)
    pdf.set_fill_color(241, 245, 249)
    pdf.set_text_color(30, 41, 59)
    pdf.set_x(10)
    for header, width in zip(headers, col_widths):
        pdf.cell(width, 7, _fit_cell_text(pdf, _pdf_safe(header), width), border=1, fill=True)
    pdf.ln(7)

    pdf.set_font("Helvetica", "", 9)
    for row in rows:
        pdf.set_x(10)
        for value, width in zip(row, col_widths):
            pdf.cell(width, 7, _fit_cell_text(pdf, _pdf_safe(value), width), border=1)
        pdf.ln(7)


def generate_pdf_report(
    industry, company_name, time_horizon, result, ai_summary, recommendations,
    commodity_data, shipping_status, logistics_result, by_country,
    critical_alerts, high_risk_suppliers, disruption_band,
    single_source_dependency, compliance_results,
    shipment_delays, port_congestion, on_time_rate,
    top_country_name, disaster_alert, pol_reg_score, weather_alert, conflict_alert,
):
    """Returns the .pdf file as bytes, ready for st.download_button.

    Renders the same commodity/sourcing charts shown on the website via kaleido
    (Plotly's static-image export backend) and embeds them as PNGs.
    """
    pdf = FPDF()
    pdf.add_page()
    pdf.set_auto_page_break(auto=True, margin=15)
    epw = pdf.epw  # effective page width, used instead of 0 to avoid fpdf2's
    # "0 = extend to right margin from current x" ambiguity, which can shrink
    # to near-zero width if x wasn't reset to the left margin beforehand.

    pdf.set_font("Helvetica", "B", 20)
    pdf.set_x(10)
    pdf.cell(epw, 10, "Supply Chain Risk Monitor", new_x="LMARGIN", new_y="NEXT")

    pdf.set_font("Helvetica", "", 11)
    pdf.set_text_color(100, 100, 100)
    subtitle = f"{industry}"
    if company_name:
        subtitle += f"  |  {company_name}"
    subtitle += f"  |  {time_horizon} horizon  |  Generated {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    pdf.set_x(10)
    pdf.cell(epw, 8, _pdf_safe(subtitle), new_x="LMARGIN", new_y="NEXT")
    pdf.ln(4)

    pdf.set_text_color(30, 41, 59)
    pdf.set_font("Helvetica", "B", 15)
    pdf.set_x(10)
    pdf.cell(epw, 9, "Executive Summary", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(2)

    # Overall score box
    color = RISK_COLOR_RGB.get(result["label"], (100, 100, 100))
    pdf.set_fill_color(*color)
    pdf.rect(10, pdf.get_y(), 190, 22, style="F")
    pdf.set_text_color(255, 255, 255)
    pdf.set_font("Helvetica", "B", 16)
    pdf.set_xy(15, pdf.get_y() + 5)
    pdf.cell(epw - 5, 10, f"Overall Risk Score: {result['total']} / 100  -  {result['label']}", new_x="LMARGIN", new_y="NEXT")
    pdf.set_x(10)
    pdf.ln(15)

    pdf.set_font("Helvetica", "B", 12)
    pdf.set_x(10)
    pdf.cell(epw, 7, "Critical Risk Alerts", new_x="LMARGIN", new_y="NEXT")
    pdf.set_font("Helvetica", "", 10)
    if critical_alerts:
        for label, score in critical_alerts:
            pdf.set_x(10)
            pdf.multi_cell(epw, 6, _pdf_safe(f"- {label}: {score} ({_risk_label(score)})"), new_x="LMARGIN", new_y="NEXT")
    else:
        pdf.set_x(10)
        pdf.multi_cell(epw, 6, "None - all categories within moderate range or better.", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(2)

    pdf.set_font("Helvetica", "B", 12)
    pdf.set_x(10)
    pdf.cell(epw, 7, "High-Risk Suppliers", new_x="LMARGIN", new_y="NEXT")
    pdf.set_font("Helvetica", "", 10)
    if high_risk_suppliers:
        for s in high_risk_suppliers:
            pdf.set_x(10)
            pdf.multi_cell(
                epw, 6,
                _pdf_safe(f"- {s['name']} ({s['detail']}): {s['risk']} ({s['label']})"),
                new_x="LMARGIN", new_y="NEXT",
            )
    else:
        pdf.set_x(10)
        pdf.multi_cell(epw, 6, "None - all sourcing within acceptable risk range.", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(2)

    pdf.set_font("Helvetica", "B", 12)
    pdf.set_x(10)
    pdf.cell(epw, 7, "Disruption Probability", new_x="LMARGIN", new_y="NEXT")
    pdf.set_font("Helvetica", "", 10)
    pdf.set_x(10)
    pdf.multi_cell(epw, 6, _pdf_safe(disruption_band[0]), new_x="LMARGIN", new_y="NEXT")
    pdf.ln(4)

    # Sub-score breakdown
    pdf.set_text_color(30, 41, 59)
    pdf.set_font("Helvetica", "B", 13)
    pdf.set_x(10)
    pdf.cell(epw, 8, "Risk Breakdown", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(1)

    for key, value in result["sub_scores"].items():
        label = CATEGORY_LABELS.get(key, key)
        risk_text = _risk_label(value)
        box_color = RISK_COLOR_RGB.get(risk_text, (100, 100, 100))
        y = pdf.get_y()
        pdf.set_fill_color(*box_color)
        pdf.rect(10, y, 4, 8, style="F")
        pdf.set_text_color(30, 41, 59)
        pdf.set_font("Helvetica", "", 11)
        pdf.set_xy(17, y)
        pdf.cell(65, 8, label)
        pdf.set_xy(85, y)
        pdf.set_font("Helvetica", "B", 11)
        pdf.cell(25, 8, f"{value}")
        pdf.set_font("Helvetica", "", 10)
        pdf.set_text_color(*box_color)
        pdf.cell(47, 8, risk_text)
        pdf.set_text_color(120, 120, 120)
        pdf.cell(35, 8, f"Weight: {result['weights'][key]:.0%}")
        pdf.set_text_color(30, 41, 59)
        pdf.set_x(10)
        pdf.ln(9)

    pdf.ln(4)
    pdf.set_font("Helvetica", "B", 15)
    pdf.set_x(10)
    pdf.cell(epw, 9, "Supplier Risk", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(2)

    for metric_label, value_text in [
        (row["Metric"], row["Value"])
        for row in _supplier_risk_rows(result, single_source_dependency, compliance_results)
    ]:
        pdf.set_font("Helvetica", "B", 11)
        pdf.set_x(10)
        pdf.multi_cell(epw, 6, _pdf_safe(metric_label), new_x="LMARGIN", new_y="NEXT")
        pdf.set_font("Helvetica", "", 10)
        pdf.set_x(10)
        pdf.multi_cell(epw, 6, _pdf_safe(value_text), new_x="LMARGIN", new_y="NEXT")
        pdf.ln(2)

    pdf.ln(2)
    pdf.set_font("Helvetica", "B", 15)
    pdf.set_x(10)
    pdf.cell(epw, 9, "Logistics Risk", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(2)

    for metric_label, value_text in [
        (row["Metric"], row["Value"])
        for row in _logistics_risk_rows(result, shipment_delays, port_congestion, on_time_rate)
    ]:
        pdf.set_font("Helvetica", "B", 11)
        pdf.set_x(10)
        pdf.multi_cell(epw, 6, _pdf_safe(metric_label), new_x="LMARGIN", new_y="NEXT")
        pdf.set_font("Helvetica", "", 10)
        pdf.set_x(10)
        pdf.multi_cell(epw, 6, _pdf_safe(value_text), new_x="LMARGIN", new_y="NEXT")
        pdf.ln(2)

    pdf.ln(2)
    pdf.set_font("Helvetica", "B", 15)
    pdf.set_x(10)
    pdf.cell(epw, 9, "Geographic & External Risk", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(2)

    for metric_label, value_text in [
        (row["Metric"], row["Value"])
        for row in _geo_external_rows(top_country_name, disaster_alert, pol_reg_score, weather_alert, conflict_alert)
    ]:
        pdf.set_font("Helvetica", "B", 11)
        pdf.set_x(10)
        pdf.multi_cell(epw, 6, _pdf_safe(metric_label), new_x="LMARGIN", new_y="NEXT")
        pdf.set_font("Helvetica", "", 10)
        pdf.set_x(10)
        pdf.multi_cell(epw, 6, _pdf_safe(value_text), new_x="LMARGIN", new_y="NEXT")
        pdf.ln(2)

    pdf.ln(2)
    pdf.set_font("Helvetica", "B", 13)
    pdf.set_x(10)
    pdf.cell(epw, 8, "AI Risk Brief", new_x="LMARGIN", new_y="NEXT")
    pdf.set_font("Helvetica", "", 10)
    pdf.set_x(10)
    pdf.multi_cell(epw, 6, _pdf_safe(ai_summary), new_x="LMARGIN", new_y="NEXT")
    pdf.ln(4)

    pdf.set_font("Helvetica", "B", 13)
    pdf.set_x(10)
    pdf.cell(epw, 8, "Recommended Actions", new_x="LMARGIN", new_y="NEXT")
    for i, rec in enumerate(recommendations, 1):
        pdf.set_font("Helvetica", "B", 11)
        pdf.set_x(10)
        pdf.multi_cell(epw, 6, _pdf_safe(f"{i}. {rec['title']} (Priority: {rec['priority']})"), new_x="LMARGIN", new_y="NEXT")
        pdf.set_font("Helvetica", "", 10)
        pdf.set_x(10)
        pdf.multi_cell(epw, 6, _pdf_safe(rec["detail"]), new_x="LMARGIN", new_y="NEXT")
        pdf.ln(2)

    # --- Detailed Analysis ---
    pdf.ln(4)
    pdf.set_text_color(30, 41, 59)
    pdf.set_font("Helvetica", "B", 15)
    pdf.set_x(10)
    pdf.cell(epw, 9, "Detailed Analysis", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(2)

    pdf.set_font("Helvetica", "B", 13)
    pdf.set_x(10)
    pdf.cell(epw, 8, "Commodity Price Trends", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(2)
    chart_width_px, chart_height_px = CHART_PNG_DIMS
    for name, history in commodity_data.items():
        if len(history) < 2:
            continue
        png_bytes = _fig_to_png(build_commodity_chart(name, history), chart_width_px, chart_height_px)
        _place_image(pdf, epw, png_bytes, chart_width_px, chart_height_px)

    commodity_rows = _commodity_rows(commodity_data)
    _pdf_table(
        pdf, epw,
        ["Commodity", "Latest Value", "% Change (period)"],
        [[r["Commodity"], r["Latest Value"], r["% Change (period)"]] for r in commodity_rows],
        [0.5, 0.25, 0.25],
    )
    pdf.ln(6)

    pdf.set_font("Helvetica", "B", 13)
    pdf.set_x(10)
    pdf.cell(epw, 8, "Shipping Route Status", new_x="LMARGIN", new_y="NEXT")
    shipping_rows = _shipping_rows(shipping_status, logistics_result)
    _pdf_table(
        pdf, epw,
        ["Route", "Status", "Delay", "Cost +%", "Risk", "News Signal"],
        [[r["Route"], r["Status"], r["Delay (days)"], r["Cost Premium (%)"], r["Risk Score"], r["News Signal"]] for r in shipping_rows],
        [0.32, 0.15, 0.1, 0.13, 0.1, 0.2],
    )
    pdf.ln(2)
    pdf.set_font("Helvetica", "B", 10)
    pdf.set_x(10)
    pdf.cell(epw, 6, f"Overall Logistics Risk: {logistics_result['score']}", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(6)

    pdf.set_font("Helvetica", "B", 13)
    pdf.set_x(10)
    pdf.cell(epw, 8, "Sourcing Risk Map", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(2)
    map_width_px, map_height_px = MAP_PNG_DIMS
    map_png_bytes = _fig_to_png(build_geo_choropleth(by_country), map_width_px, map_height_px)
    _place_image(pdf, epw, map_png_bytes, map_width_px, map_height_px)

    pdf.set_font("Helvetica", "B", 13)
    pdf.set_x(10)
    pdf.cell(epw, 8, "Sourcing Risk Breakdown", new_x="LMARGIN", new_y="NEXT")
    sourcing_rows = _sourcing_rows(by_country)
    _pdf_table(
        pdf, epw,
        ["Country", "Code", "% Sourcing", "Product", "Risk"],
        [[r["Country"], r["Code"], r["% of Sourcing"], r["Product"], r["Risk Score"]] for r in sourcing_rows],
        [0.25, 0.1, 0.13, 0.37, 0.15],
    )

    return bytes(pdf.output())
